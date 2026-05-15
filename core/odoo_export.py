from __future__ import annotations

import io
from dataclasses import asdict
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .payroll_excel import PayrollEmployee
from .payroll_pdf import PayslipPage
from .utils import decimal_to_float, is_zero, money, safe_account_465


DEFAULTS = {
    "account_salary": "64000000",
    "account_ss_company": "64200100",
    "account_irpf": "47510000",
    "account_ss_creditor": "47600012",
    "account_embargo": "46599999",
    "account_especie_autonomo": "46599998",
    "account_indemniz_inss": "47100000",
    "tax_grid_salary": "mod111[02]",
    "tax_grid_irpf": "mod111[03]",
}


def build_employee_rows(employees: list[PayrollEmployee], pdf_pages: list[PayslipPage]) -> list[dict[str, Any]]:
    pdf_by_worker = {p.worker_number: p for p in pdf_pages if p.worker_number}
    rows: list[dict[str, Any]] = []
    for emp in employees:
        pdf = pdf_by_worker.get(emp.worker_number)
        rows.append(
            {
                "incluir": "SI",
                "n_trabajador": emp.worker_number,
                "dni": pdf.dni if pdf else "",
                "nombre_excel": emp.name,
                "nombre_pdf": pdf.employee_name if pdf else "",
                "centro": emp.center or "",
                "cuenta_sueldos": DEFAULTS["account_salary"],
                "cuenta_ss_empresa": DEFAULTS["account_ss_company"],
                "cuenta_remuneraciones": safe_account_465(emp.worker_number),
                "cuenta_irpf": DEFAULTS["account_irpf"],
                "cuenta_ss_acreedora": DEFAULTS["account_ss_creditor"],
                "cuenta_embargo": DEFAULTS["account_embargo"],
                "cuenta_especie_autonomo": DEFAULTS["account_especie_autonomo"],
                "cuenta_indemniz_inss": DEFAULTS["account_indemniz_inss"],
                "partner_odoo": emp.name,
                "telefono": "",
                "tax_grid_sueldo": DEFAULTS["tax_grid_salary"],
                "tax_grid_irpf": DEFAULTS["tax_grid_irpf"],
                "total_bruto": decimal_to_float(emp.concepts.get("total_bruto", Decimal("0"))),
                "total_neto": decimal_to_float(emp.concepts.get("total_neto", Decimal("0"))),
                "descuento_ss": decimal_to_float(emp.concepts.get("descuento_ss", Decimal("0"))),
                "descuento_irpf": decimal_to_float(emp.concepts.get("descuento_irpf", Decimal("0"))),
                "ss_empresa": decimal_to_float(emp.concepts.get("ss_empresa", Decimal("0"))),
                "coste_tc1": decimal_to_float(emp.concepts.get("coste_tc1", Decimal("0"))),
                "embargo_juzgado": decimal_to_float(emp.concepts.get("embargo_juzgado", Decimal("0"))),
                "valores_especie_autonomo": decimal_to_float(emp.concepts.get("valores_especie_autonomo", Decimal("0"))),
                "total_indemniz_inss": decimal_to_float(emp.concepts.get("total_indemniz_inss", Decimal("0"))),
            }
        )
    return rows


def mapping_template_xlsx(employee_rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "mapeo_empleados"
    headers = [
        "incluir",
        "n_trabajador",
        "dni",
        "nombre_excel",
        "nombre_pdf",
        "centro",
        "cuenta_sueldos",
        "cuenta_ss_empresa",
        "cuenta_remuneraciones",
        "cuenta_irpf",
        "cuenta_ss_acreedora",
        "cuenta_embargo",
        "cuenta_especie_autonomo",
        "cuenta_indemniz_inss",
        "partner_odoo",
        "telefono",
        "tax_grid_sueldo",
        "tax_grid_irpf",
        "total_bruto",
        "total_neto",
        "descuento_ss",
        "descuento_irpf",
        "ss_empresa",
        "coste_tc1",
        "embargo_juzgado",
        "valores_especie_autonomo",
        "total_indemniz_inss",
    ]
    ws.append(headers)
    for row in employee_rows:
        ws.append([row.get(h, "") for h in headers])
    _format_sheet(ws)

    info = wb.create_sheet("instrucciones")
    info.append(["Campo", "Uso"])
    info.append(["cuenta_sueldos", "Cuenta de gasto 640 que debe revisarse por trabajador/categoría."])
    info.append(["cuenta_remuneraciones", "Cuenta 465 de remuneraciones pendientes. Se precarga con 46510 + nº trabajador."])
    info.append(["cuenta_ss_acreedora", "Cuenta 476 para el crédito agregado del TC1."])
    info.append(["cuenta_especie_autonomo", "Cuenta para compensar valores en especie/autónomo que figuran como deducción."])
    info.append(["cuenta_indemniz_inss", "Cuenta para prestaciones/indemnizaciones INSS incluidas en devengos, si procede."])
    info.append(["telefono", "Telefono movil para envio de enlace de firma por WhatsApp (formato internacional recomendado)."])
    info.append(["tax_grid_sueldo", "Etiqueta/cuadrícula fiscal para bases sujetas a retención si se usa modelo 111 en Odoo."])
    info.append(["tax_grid_irpf", "Etiqueta/cuadrícula fiscal para la línea acreedora de retenciones."])
    _format_sheet(info)
    return _save_bytes(wb)


def read_mapping_xlsx(xlsx_path: str | Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(filename=xlsx_path, data_only=True)
    ws = wb["mapeo_empleados"] if "mapeo_empleados" in wb.sheetnames else wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    result: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        worker = str(record.get("n_trabajador") or "").strip()
        if worker:
            result[worker] = record
    return result


def generate_odoo_import_xlsx(
    employees: list[PayrollEmployee],
    employee_rows: list[dict[str, Any]],
    month: int,
    year: int,
    journal: str,
    date: str,
    reference: str,
    aggregate_ss: bool = True,
) -> tuple[bytes, dict[str, Any]]:
    mapping = {str(row["n_trabajador"]): row for row in employee_rows}
    move_external_id = f"nomina_{year}_{month:02d}"
    lines: list[dict[str, Any]] = []
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")
    total_tc1 = Decimal("0.00")

    for emp in employees:
        map_row = mapping.get(emp.worker_number, {})
        if str(map_row.get("incluir", "SI")).upper() not in ("SI", "S", "YES", "TRUE", "1"):
            continue
        label_base = f"NÓMINA {month:02d}/{year} {emp.worker_number} {emp.name}"
        gross = money(emp.concepts.get("total_bruto", Decimal("0")))
        net = money(emp.concepts.get("total_neto", Decimal("0")))
        irpf = money(emp.concepts.get("descuento_irpf", Decimal("0")))
        ss_company = money(emp.concepts.get("ss_empresa", Decimal("0")))
        tc1 = money(emp.concepts.get("coste_tc1", Decimal("0")))
        embargo = money(emp.concepts.get("embargo_juzgado", Decimal("0")))
        especie = money(emp.concepts.get("valores_especie_autonomo", Decimal("0")))
        indemniz_inss = money(emp.concepts.get("total_indemniz_inss", Decimal("0")))
        total_tc1 += tc1

        if not is_zero(gross):
            lines.append(_line(move_external_id, date, journal, reference, map_row.get("cuenta_sueldos"), map_row.get("partner_odoo"), label_base, gross, 0, map_row.get("tax_grid_sueldo")))
            total_debit += gross
        if not is_zero(ss_company):
            lines.append(_line(move_external_id, date, journal, reference, map_row.get("cuenta_ss_empresa"), map_row.get("partner_odoo"), f"SEGURIDAD SOCIAL EMPRESA {emp.worker_number} {emp.name}", ss_company, 0, ""))
            total_debit += ss_company
        if not is_zero(net):
            lines.append(_line(move_external_id, date, journal, reference, map_row.get("cuenta_remuneraciones"), map_row.get("partner_odoo"), f"LÍQUIDO NÓMINA {emp.worker_number} {emp.name}", 0, net, ""))
            total_credit += net
        if not is_zero(irpf):
            lines.append(_line(move_external_id, date, journal, reference, map_row.get("cuenta_irpf"), map_row.get("partner_odoo"), f"IRPF NÓMINA {emp.worker_number} {emp.name}", 0, irpf, map_row.get("tax_grid_irpf")))
            total_credit += irpf
        if not is_zero(embargo):
            lines.append(_line(move_external_id, date, journal, reference, map_row.get("cuenta_embargo"), map_row.get("partner_odoo"), f"EMBARGO JUZGADO {emp.worker_number} {emp.name}", 0, embargo, ""))
            total_credit += embargo
        if not is_zero(especie):
            lines.append(_line(move_external_id, date, journal, reference, map_row.get("cuenta_especie_autonomo"), map_row.get("partner_odoo"), f"VALORES EN ESPECIE/AUTÓNOMO {emp.worker_number} {emp.name}", 0, especie, ""))
            total_credit += especie
        if not is_zero(indemniz_inss):
            lines.append(_line(move_external_id, date, journal, reference, map_row.get("cuenta_indemniz_inss"), map_row.get("partner_odoo"), f"PRESTACIONES/INDEMNIZACIONES INSS {emp.worker_number} {emp.name}", 0, indemniz_inss, ""))
            total_credit += indemniz_inss

        if not aggregate_ss and not is_zero(tc1):
            lines.append(_line(move_external_id, date, journal, reference, map_row.get("cuenta_ss_acreedora"), map_row.get("partner_odoo"), f"TC1 {emp.worker_number} {emp.name}", 0, tc1, ""))
            total_credit += tc1

    if aggregate_ss and not is_zero(total_tc1):
        # Use the first mapping row to retrieve the SS creditor account.
        first_map = next(iter(mapping.values()), {}) if mapping else {}
        lines.insert(0, _line(move_external_id, date, journal, reference, first_map.get("cuenta_ss_acreedora", DEFAULTS["account_ss_creditor"]), "", f"TC1 NÓMINA {month:02d}/{year}", 0, total_tc1, ""))
        total_credit += total_tc1

    balance = money(total_debit - total_credit)
    summary = {
        "line_count": len(lines),
        "total_debit": decimal_to_float(total_debit),
        "total_credit": decimal_to_float(total_credit),
        "balance": decimal_to_float(balance),
        "is_balanced": is_zero(balance),
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "odoo_import"
    headers = [
        "id",
        "move_type",
        "date",
        "journal_id",
        "ref",
        "line_ids/account_id",
        "line_ids/partner_id",
        "line_ids/name",
        "line_ids/debit",
        "line_ids/credit",
        "line_ids/tax_tag_ids",
        "line_ids/tax_ids",
        "line_ids/analytic_account_id",
    ]
    ws.append(headers)
    for line in lines:
        ws.append([line.get(h, "") for h in headers])
    _format_sheet(ws)

    s = wb.create_sheet("resumen")
    s.append(["Métrica", "Valor"])
    for key, value in summary.items():
        s.append([key, value])
    s.append([])
    s.append(["Nota", "Si el balance no es 0, revisar embargos u otras deducciones no mapeadas."])
    _format_sheet(s)
    return _save_bytes(wb), summary


def _line(external_id, date, journal, ref, account, partner, label, debit, credit, tax_grid):
    return {
        "id": external_id,
        "move_type": "entry",
        "date": date,
        "journal_id": journal or "NOMINAS",
        "ref": ref,
        "line_ids/account_id": account or "",
        "line_ids/partner_id": partner or "",
        "line_ids/name": label,
        "line_ids/debit": decimal_to_float(money(debit)),
        "line_ids/credit": decimal_to_float(money(credit)),
        "line_ids/tax_tag_ids": tax_grid or "",
        "line_ids/tax_ids": "",
        "line_ids/analytic_account_id": "",
    }


def _format_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for column in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_len + 2, 12), 45)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _save_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
