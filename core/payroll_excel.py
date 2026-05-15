from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .utils import normalize_concept, parse_decimal


@dataclass
class PayrollEmployee:
    worker_number: str
    name: str
    center: str | None = None
    concepts: dict[str, Any] = field(default_factory=dict)

    def get(self, concept_name: str):
        return self.concepts.get(concept_name, 0)


EMPLOYEE_HEADER_RE = re.compile(r"^\s*(\d+)\s*-\s*(.+?)\s*$")
PERIOD_RE = re.compile(r"(\d{2})-(\d{4})\s+a\s+(\d{2})-(\d{4})")

# Canonical concept names used later in the accounting export.
CONCEPT_ALIASES = {
    "TOTAL BRUTO": "total_bruto",
    "TOTAL NETO": "total_neto",
    "TOTAL DESCUENTO S.S.": "descuento_ss",
    "TOTAL DESCUENTO IRPF": "descuento_irpf",
    "SEG. SOCIAL EMPRESA": "ss_empresa",
    "COSTE TC1": "coste_tc1",
    "EMBARGO JUZGADO (NOMINA)": "embargo_juzgado",
    "TOTAL INDEMNIZ INSS": "total_indemniz_inss",
    "VALORES EN ESPECIE AUTONOMO": "valores_especie_autonomo",
    "BASE I.R.P.F.": "base_irpf",
}


def parse_payroll_excel(xlsx_path: str | Path) -> tuple[list[PayrollEmployee], dict[str, Any]]:
    wb = load_workbook(filename=xlsx_path, data_only=True)
    ws = wb.active

    metadata: dict[str, Any] = {
        "period_month": None,
        "period_year": None,
        "company": None,
        "centers": [],
    }

    for row in ws.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str):
                match = PERIOD_RE.search(value)
                if match and metadata["period_month"] is None:
                    metadata["period_month"] = int(match.group(1))
                    metadata["period_year"] = int(match.group(2))
                if value.strip().startswith("25 - ") and metadata["company"] is None:
                    metadata["company"] = value.strip()

    employees: dict[str, PayrollEmployee] = {}
    max_row = ws.max_row
    row = 1

    while row <= max_row:
        first_value = ws.cell(row=row, column=1).value
        if normalize_concept(first_value) != "CONCEPTO SALARIAL":
            row += 1
            continue

        center = _find_center_in_previous_rows(ws, row)
        if center and center not in metadata["centers"]:
            metadata["centers"].append(center)

        employee_cols: list[tuple[int, str, str]] = []
        for col in range(2, ws.max_column + 1):
            header = ws.cell(row=row, column=col).value
            if not isinstance(header, str):
                continue
            if header.strip().upper().startswith("TOTAL"):
                continue
            match = EMPLOYEE_HEADER_RE.match(header)
            if match:
                employee_cols.append((col, match.group(1), match.group(2).strip()))

        row += 1
        while row <= max_row:
            concept_raw = ws.cell(row=row, column=1).value
            normalized = normalize_concept(concept_raw)
            if normalized in ("CONCEPTO SALARIAL", "") or normalized.startswith("LISTADO DE DETALLE"):
                break
            canonical_key = CONCEPT_ALIASES.get(normalized)
            if concept_raw is not None:
                for col, worker_number, employee_name in employee_cols:
                    emp = employees.get(worker_number)
                    if emp is None:
                        emp = PayrollEmployee(worker_number=worker_number, name=employee_name, center=center)
                        employees[worker_number] = emp
                    if not emp.center and center:
                        emp.center = center
                    value = ws.cell(row=row, column=col).value
                    # Keep both canonical concepts and original concepts to allow later extensions.
                    if canonical_key:
                        emp.concepts[canonical_key] = parse_decimal(value)
                    emp.concepts[str(concept_raw).strip()] = parse_decimal(value)
            row += 1
        # Do not increment row again; outer loop should inspect current break row.
    return sorted(employees.values(), key=lambda e: int(e.worker_number)), metadata


def _find_center_in_previous_rows(ws, header_row: int) -> str | None:
    for r in range(max(1, header_row - 4), header_row):
        for c in range(1, ws.max_column + 1):
            value = ws.cell(row=r, column=c).value
            if isinstance(value, str) and value.strip().startswith("Centro:"):
                next_value = ws.cell(row=r, column=c + 1).value
                return str(next_value).strip() if next_value else None
    return None
